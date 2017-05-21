# coding: utf8

import time
import xlsxwriter
import datetime

from impeller.c_workbook import Workbook

ROWS = 10**4
COLS = 10


def test_xlsxwriter(workbook):
    workbook.set_properties({"author": "Емельянов Дмитрий", "company": "Нету компании"})
    workbook.set_custom_property("Кастом юникод дата", datetime.datetime.now())
    ws = workbook.add_worksheet("Юникод шит!")

    start = time.clock()

    bold_format = workbook.add_format()
    bold_format.set_bold()

    zebra_format = workbook.add_format()
    zebra_format.set_bg_color('#d3d3d3')
    zebra_format.set_italic()

    for i in range(ROWS):
        for j in range(COLS):
            # row, col, data
            if i % 2 == 0:
                ws.write(i, j, i + j, zebra_format if i else bold_format)
            else:
                ws.write(i, j, i + j, None)

    ws.set_column(0, 2, 30)
    ws.set_column("C:E", 50)
    start_close = time.clock()
    workbook.close()
    end_close = time.clock()
    end = time.clock()
    print("Close time:", format(end_close - start_close, ".2f"), "Total:", format(end - start, ".2f"))


def test_openpyxl(wb):
    ws = wb.create_sheet("Юникод шит")

    start = time.clock()

    bold_font = Font(bold=True)
    italic_font = Font(italic=True)
    grey_background = PatternFill(bgColor='d3d3d3', fill_type='solid')

    for i in range(ROWS):
        row = []
        for j in range(COLS):
            # row, col, data
            cell_obj = WriteOnlyCell(ws=ws, value=i + j)
            if i % 2 == 0:
                if i:
                    cell_obj.font = italic_font
                    cell_obj.fill = grey_background
                else:
                    cell_obj.font = bold_font
            row.append(cell_obj)
        ws.append(row)

    ws.column_dimensions["A"].width = 30.0
    ws.column_dimensions["B"].width = 30.0
    ws.column_dimensions["C"].width = 50.0
    ws.column_dimensions["D"].width = 50.0
    ws.column_dimensions["E"].width = 50.0

    start_close = time.clock()
    wb.save("openpyxl.xlsx")
    end_close = time.clock()
    end = time.clock()
    print("Close time:", format(end_close - start_close, ".2f"), "Total:", format(end - start, ".2f"))


def test_pyexcelerate(wb):
    # Original is run_pyexcelerate_style_cheating from pyexcelerate bench
    ws = wb.new_sheet('Юникод шит')

    start = time.clock()

    bold_font = xlr_Style(font=xlr_Font(bold=True))
    zebra_style = xlr_Style(font=xlr_Font(italic=True), fill=xlr_Fill(background=xlr_Color(211, 211, 211)))

    for i in range(ROWS):
        for j in range(COLS):
            ws.set_cell_value(i + 1, j + 1, i + j)
            if i % 2 == 0:
                if i:
                    ws.set_cell_style(i + 1, j + 1, zebra_style)
                else:
                    ws.set_cell_style(i + 1, j + 1, bold_font)

    start_close = time.clock()
    wb.save("pyexcelerate.xlsx")
    end_close = time.clock()
    end = time.clock()
    print("Close time:", format(end_close - start_close, ".2f"), "Total:", format(end - start, ".2f"))


def test_xlwt_xlsx(wb):
    # Max 65536 elements
    # ROWS = 6500
    # COLS = 10
    # time_multiplier = 10**5 / 6500

    ws = wb.add_sheet('Юникод шит')

    start = time.clock()

    xlwt.add_palette_colour("light_gray", 0x21)
    wb.set_colour_RGB(0x21, 211, 211, 211)

    bold_font = easyxf("font: bold on")
    zebra_style = easyxf("font: italic on; pattern: pattern solid, back-colour light_gray")

    for i in range(ROWS):
        for j in range(COLS):
            # row, col, data
            if i % 2 == 0:
                if i:
                    ws.write(i, j, i + j, style=zebra_style)
                else:
                    ws.write(i, j, i + j, style=bold_font)
            else:
                ws.write(i, j, i + j)

    start_close = time.clock()
    wb.save("xlwt.xls")
    end_close = time.clock()
    end = time.clock()
    print("Close time:", format(end_close - start_close, ".2f"), "Total:", format(end - start, ".2f"))

c_wb = Workbook("impeller.xlsx")
print("Impeller timing:")
test_xlsxwriter(c_wb)

wb = xlsxwriter.Workbook("plain_xlsxwriter.xlsx")
print("XlsxWriter timing:")
test_xlsxwriter(wb)


if ROWS * COLS < 65536:
    import xlwt
    from xlwt import Workbook as xlwt_Workbook
    from xlwt import easyxf
    xlwt_wb = xlwt_Workbook()
    print("xlwt timing:")
    test_xlwt_xlsx(xlwt_wb)


from pyexcelerate import Workbook as PyExcelerate_Workbook
from pyexcelerate.Color import Color as xlr_Color
from pyexcelerate.Style import Style as xlr_Style
from pyexcelerate.Font import Font as xlr_Font
from pyexcelerate.Fill import Fill as xlr_Fill
excelr_wb = PyExcelerate_Workbook()
print("pyexcelerate timing")
test_pyexcelerate(excelr_wb)


from openpyxl import Workbook as PXL_Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.writer.write_only import WriteOnlyCell
pxl_wb = PXL_Workbook(write_only=True)
print("openpyxl timing:")
test_openpyxl(pxl_wb)
